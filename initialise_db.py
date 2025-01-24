import openpyxl
import sqlite3 as sql
from time import sleep
from db_tools import start_db, add_to_db
from add_book import add_books

# Sleep before sending message.
def sleep_with_message(msg):
    sleep(1)
    print(msg)

# Ask the user for their seeding method of choice, verify input.
def ask_seed_method():
    seed_method = int(input("Would you like to...\n"
                            "1: Start with an empty library.\n"
                            "2: Start by adding a few books.\n"
                            "3: Import library from excel file.\n"
                            "(Type a number and hit enter to begin.)\n"
                            ))
    if seed_method in (1, 2, 3):
        return seed_method
    else:
        sleep_with_message("Incorrect input. Trying again...")
        ask_seed_method()

# Ask the user for the name of their excel file, get the data, add to database.
def import_from_excel():
    sleep_with_message("Please ensure your excel file is in the same folder as this program, then..")
    excel_name = input("Enter file name (everything before the dot): ")

    try:
        data = openpyxl.load_workbook("{}.xlsx".format(excel_name))
    except:
        sleep_with_message("File not found. Trying again...")
        import_from_excel()

    # Gather the book titles from excel.
    books_list = []
    data_r = data.active
    for row in range(0, data_r.max_row):
        for col in data_r.iter_cols(1, data_r.max_column):
            books_list.append(col[row].value)

    add_to_db(books_list)

# Start database, call relevant seeding method, inform user of next steps.
if __name__ == "__main__":
    seed_method = ask_seed_method()
    start_db()
    match seed_method:
        case 1:
            sleep_with_message("Database created successfully.")
            sleep_with_message(
                              "Next steps:\n"
                              "• Add some books by running add_book.py\n"
                              "• Once you have added some books, get a random book by running get_book.py\n"
                               )
        case 2:
            add_books()
            sleep_with_message("Your books have been added to the database.")
            sleep_with_message(
                              "Next steps:\n"
                              "• Add more books by running add_book.py\n"
                              "• Get a random book by running get_book.py\n"
                               )
        case 3:
            sleep(1)
            import_from_excel()
            sleep_with_message("Your books have been imported into the database.")
            sleep_with_message(
                              "Next steps:\n"
                              "• Add more books by running add_book.py\n"
                              "• Get a random book by running get_book.py\n"
                               )

    sleep(7)    
